import openpyxl
import pandas as pd
from scipy.spatial import distance_matrix
from docs.cell import Cell
from docs.utils import calculate_distance_in_same_block

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("mezanin_cells_by_floors.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

cells = []

# Iterate the loop to read the cell values
for row_index in range(0, dataframe1.max_row):
    x = 0
    if row_index % 10 == 0:
        x += 1
    for col_index in range(0, dataframe1.max_column):
        code = dataframe1.cell(row_index + 1, col_index + 1).value
        if code is not None:
            y = col_index + 1
            cells.append(Cell(x, y, 1, code))

print(len(cells))

distance_data = []
codes = []
for i in range(len(cells)):
    codes.append(cells[i].code)
    tmp_dist_data = []
    for j in range(len(cells)):
        if i != j:
            tmp_dist_data.append(calculate_distance_in_same_block(cells[i], cells[j], 99))
        else:
            tmp_dist_data.append(0)
    distance_data.append(tmp_dist_data)
    print(round(i / 27955 * 100), '%')

df = pd.DataFrame(distance_data, columns=codes, index=codes)

df.to_csv("distance_matrix.csv", index=True)